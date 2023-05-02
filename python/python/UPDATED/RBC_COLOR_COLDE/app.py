from datetime import datetime
import sys
import collections
import pandas as pd
import xml.etree.ElementTree as et
from openpyxl.styles import PatternFill
pd.options.mode.chained_assignment = None
pd.set_option("display.max_colwidth", 10000)
print("Program executing Please wait!")
start=datetime.now()
ctm_file1,ctm_file2=sys.argv[1],sys.argv[2]

class COLOR_CODE:
    def __init__(self):
        pass
    @staticmethod
    def ctm_details(ctm_file):
        parsexml=et.parse(ctm_file)
        root=parsexml.getroot()
        xml=[]
        for r in root:
            Folder_name=r.attrib.get('FOLDER_NAME',"")
            Parent_folder=r.attrib.get('PARENT_FOLDER',Folder_name)
            RBC=[" ".join(a.attrib.values()) for a in r.findall('RULE_BASED_CALENDAR')]
            xml.append([Parent_folder,Folder_name,r.tag,'',RBC])
            JOB=r.findall('JOB')
            for j in JOB:
                var=j.attrib
                PARENT_FOLDER,JOBNAME=var.get('PARENT_FOLDER',""),var.get("JOBNAME","")
                RBC=[a.attrib.get('NAME','') for a in j.findall('RULE_BASED_CALENDARS')]
                xml.append([PARENT_FOLDER,'','JOB',JOBNAME,RBC])
            for r1 in r.findall("./"):  
                if r1.tag in ['FOLDER','SMART_FOLDER','SUB_FOLDER']:
                    Folder_name,Parent_folder=r1.attrib.get('JOBNAME',""),r1.attrib.get('PARENT_FOLDER',"")
                    RBC=[a.attrib.get('NAME','') for a in r.findall('RULE_BASED_CALENDARS')]
                    xml.append([Parent_folder,Folder_name,r1.tag,'',RBC])
                JOB=r1.findall('JOB')
                for j in JOB:
                    var=j.attrib
                    PARENT_FOLDER,JOBNAME=var.get('PARENT_FOLDER',""),var.get("JOBNAME","")
                    RBC=[a.attrib.get('NAME','') for a in j.findall('RULE_BASED_CALENDARS')]
                    xml.append([PARENT_FOLDER,'','JOB',JOBNAME,RBC])
                for r2 in r1.findall("./"):
                    if r2.tag in ['FOLDER','SMART_FOLDER','SUB_FOLDER']:
                        Folder_name,Parent_folder=r2.attrib.get('JOBNAME',""),r2.attrib.get('PARENT_FOLDER',"")
                        RBC=[a.attrib.get('NAME','') for a in r.findall('RULE_BASED_CALENDARS')]
                        xml.append([Parent_folder,Folder_name,r2.tag,'',RBC])
                    JOB=r2.findall('JOB')
                    for j in JOB:
                        var=j.attrib
                        PARENT_FOLDER,JOBNAME=var.get('PARENT_FOLDER',""),var.get("JOBNAME","")
                        RBC=[a.attrib.get('NAME','') for a in j.findall('RULE_BASED_CALENDARS')]
                        xml.append([PARENT_FOLDER,'','JOB',JOBNAME,RBC])     
        return xml
    @staticmethod
    def compare(header,df1,df2):
        data=[list(df1.columns)]
        for c_index,col in enumerate(list(df1[header])):
            rows=df2[df2[header]==col].index.tolist() 
            if len(rows)>0:
                d1=list(df1.iloc[c_index])
                d2=list(df2.iloc[rows[0]])
                local=[]
                def new1(matches,addition,deletion,duplicates):
                    if len(addition)>1 and len(deletion)>1:
                        new=matches+addition+deletion
                        if len(duplicates)>1:
                            new +=duplicates  
                        return new
                    elif len(addition)>1 and len(deletion)<2:
                        new=matches+addition
                        if len(duplicates)>1:
                            new +=duplicates 
                        return new
                    elif len(addition)<2 and len(deletion)>1:
                        new=matches+deletion
                        if len(duplicates)>1:
                            new +=duplicates
                        return new
                    else:
                        if len(duplicates)>1:
                            return matches+duplicates 
                        else:
                            return matches
                    
                for i,v in enumerate(d2):
                    matches=[]
                    addition=['addition']
                    deletion=['deletion']
                    duplicates=["duplicates"]
                    if str(d1[i])==str(d2[i]) and type(v) != list:
                        local.append(d2[i])
                    elif str(d1[i])!=str(d2[i]) and type(v) != list:
                        local.append("old: "+str(d1[i])+" \nnew: "+str(d2[i]))
                    else:
                        if len(d2[i])==len(d1[i]):
                            d2_duplicate=[item for item, count in collections.Counter(d2[i]).items() if count > 1]
                            d1_duplicate=[item for item, count in collections.Counter(d1[i]).items() if count > 1]
                            dd=list(set(d1_duplicate+d2_duplicate))
                            if len(dd)>0:
                                duplicates.append(dd)
                            for da in v:
                                if da in d1[i]:
                                    if da not in matches:
                                        matches.append(da)
                                else:
                                    if da not in addition:
                                        addition.append(da)
                            for da in d1[i]:
                                if da in v:
                                    if da not in matches:
                                        matches.append(da)
                                else:
                                    if da not in deletion:
                                        deletion.append(da)    
                        elif len(d2[i])>len(d1[i]):
                            d2_duplicate=[item for item, count in collections.Counter(d2[i]).items() if count > 1]
                            d1_duplicate=[item for item, count in collections.Counter(d1[i]).items() if count > 1]
                            dd=list(set(d1_duplicate+d2_duplicate))
                            if len(dd)>0:
                                duplicates.append(dd)
                            for da in v:
                                if da in d1[i]:
                                    if da not in matches:
                                        matches.append(da)
                                else:
                                    if da not in addition:
                                        addition.append(da)    
                        else:
                            d2_duplicate=[item for item, count in collections.Counter(d2[i]).items() if count > 1]
                            d1_duplicate=[item for item, count in collections.Counter(d1[i]).items() if count > 1]
                            dd=list(set(d1_duplicate+d2_duplicate))
                            if len(dd)>0:
                                duplicates.append(dd)
                            for da in d1[i]:
                                if da in v:
                                    if da not in matches:
                                        matches.append(da)
                                else:
                                    if da not in deletion:
                                        deletion.append(da)
                        local.append(new1(matches,addition,deletion,duplicates))
                        

                                    
                data.append(local)
        return data    
    @staticmethod
    def deleted(df1,df2,column):
        deleted_folder_name=[]
        deleted_folder_data=[]
        for i in list(df1[column]):
            if i not in list(df2[column]):
                deleted_folder_name.append(i)
                deleted_folder_data.append(list(df1.iloc[df1[df1[column]==i].index.tolist()[0]]))
        return deleted_folder_name,deleted_folder_data
    @staticmethod
    def addition(df1,df2,column):
        addition_folder_name=[]
        addition_folder_data=[]
        for i in list(df2[column]):
            if i not in list(df1[column]):
                addition_folder_name.append(i)
                addition_folder_data.append(list(df2.iloc[df2[df2[column]==i].index.tolist()[0]]))
        return addition_folder_name,addition_folder_data
    @staticmethod
    def RBC_NEW_COLUMNS(l):
        new_l=[]
        def try1(i,key):
            try:
                a=i.index(key)
            except:
                a=-1
            return a
        for i in l:
            data1=['','']
            a=try1(i,'addition')
            d=try1(i,'deletion')
            if a==-1 and d==-1:
                pass
            elif a!=-1 and d!=-1:
                data1[1]="\n".join(i[a+1:d])
                data1[0]="\n".join(i[d+1:])
            elif a!=-1 and d==-1:
                data1[1]="\n".join(i[a+1:])
            else:
                data1[0]="\n".join(i[d+1:])
            new_l.append(data1)
        return new_l

old_details=pd.DataFrame(COLOR_CODE.ctm_details(ctm_file1),columns=['PARENT_FOLDER','FOLDER_NAME','TAG','JOBNAME','RBC_DEFINITION'])
old_details['PARENT_JOBNAME']=old_details['PARENT_FOLDER']+old_details['JOBNAME']
new_details=pd.DataFrame(COLOR_CODE.ctm_details(ctm_file2),columns=['PARENT_FOLDER','FOLDER_NAME','TAG','JOBNAME','RBC_DEFINITION'])
new_details['PARENT_JOBNAME']=new_details['PARENT_FOLDER']+new_details['JOBNAME']

new_data=COLOR_CODE.compare('PARENT_JOBNAME',old_details,new_details)
fina_data= pd.DataFrame(columns=new_data[0], data=new_data[1:])


deleted_=COLOR_CODE.deleted(old_details,new_details,'PARENT_JOBNAME')

cel_deleted_Folder=[]
for i in deleted_[1]:
    fina_data.loc[len(fina_data)] = i
    c=list('ABCDE')    
    for j in c:
        cel_deleted_Folder.append(j+str(len(fina_data)+1))
        
addition_=COLOR_CODE.addition(old_details,new_details,'PARENT_JOBNAME')

cel_addition_Folder=[]
for i in addition_[1]:
    fina_data.loc[len(fina_data)] = i
    c=list('ABCDE')    
    for j in c:
        cel_addition_Folder.append(j+str(len(fina_data)+1))
rbc_changes=pd.DataFrame(COLOR_CODE.RBC_NEW_COLUMNS(list(fina_data['RBC_DEFINITION'])),columns=['OLD_RBC','NEW_RBC'])   
fina_data=pd.concat([fina_data, rbc_changes], axis=1)
fina_data['RBC_DEFINITION']=["\n".join(i) for i in fina_data['RBC_DEFINITION']]
fina_data=fina_data.drop('PARENT_JOBNAME',axis=1)
output="Chnages_"+str(datetime.now().strftime("%H%M%S"))+".xlsx"
with pd.ExcelWriter(output) as writer:
    fina_data.to_excel(writer, sheet_name='CHANGES',index=False)
import openpyxl
wb = openpyxl.load_workbook(output)
def fill_color(output,sheet,cel_deleted,cel_addition):
    dataframe=pd.read_excel(output, sheet_name=sheet)
    ws = wb[sheet]
    fill_cell6 = PatternFill(patternType='solid', fgColor='99CC32')
    for k in cel_addition:
        ws[k].fill = fill_cell6
        
    fill_cell5 = PatternFill(patternType='solid', fgColor='EE5C42')
    for j in cel_deleted:
        ws[j].fill = fill_cell5
    
    fill_cell4 = PatternFill(patternType='solid', fgColor='FFEC8B')
    fill_cell5 = PatternFill(patternType='solid', fgColor='FFD39B')
    for i,value in enumerate(list(dataframe.columns)):
        for j,v in enumerate(list(dataframe[value])):
            if 'old:' in str(v) or 'addition' in str(v) or 'deletion' in str(v) or 'duplicates' in str(v):
                if i>25:
                    c='A'+chr(i-26+65)+str(j+2)
                    ws[c].fill = fill_cell4
                else:
                    c=chr(i+65)+str(j+2)
                    ws['A'+str(j+2)].fill = fill_cell5
                    ws[c].fill = fill_cell4
fill_color(output,'CHANGES',cel_deleted_Folder,cel_addition_Folder)

wb.save(output)
# with pd.ExcelWriter("Deatails_"+str(datetime.now().strftime("%H%M%S"))+".xlsx") as w:
#     old_details.to_excel(w,sheet_name='old',index=False)
#     new_details.to_excel(w,sheet_name='new',index=False)
print(f"completed in {datetime.now()-start}")