import xml.etree.ElementTree as et
import pandas as pd
from openpyxl.styles import PatternFill
import collections
def VARIABLE1(VARIABLE,keypoint):
    L_VARIABLE=[]
    if VARIABLE:
        if 'VARIABLE' not in keypoint:
            keypoint.append('VARIABLE')
        
        for B in VARIABLE:
            s=list(B.attrib.values())
            L_VARIABLE.append(" ".join(s))   
        return L_VARIABLE
    else:
        if 'VARIABLE' not in keypoint:
            keypoint.append('VARIABLE')
        return []
def RBC1(RBC,keypoint):
    L_RBC=[]
    if RBC:
        if 'RBC' not in keypoint:
            keypoint.append('RBC')

        for a in RBC:
            s=list(a.attrib.values())
            L_RBC.append(" ".join(s))
        return L_RBC
    else:
        if 'RBC' not in keypoint:
            keypoint.append('RBC')
        return []
def CONTROL1(CONTROL,keypoint):
    L_CONTROL=[]
    if CONTROL:
        if 'CONTROL' not in keypoint:
            keypoint.append('CONTROL')

        for a in CONTROL:
            s=list(a.attrib.values())
            L_CONTROL.append(str(s[0])+' '+str(s[1]))
        return L_CONTROL
    else:
        if 'CONTROL' not in keypoint:
            keypoint.append('CONTROL')
        return '' 

def QUANTITATIVE1(QUA,keypoint):
    L_QUA=[]
    if QUA:
        if 'QUANTITATIVE' not in keypoint:
            keypoint.append('QUANTITATIVE')
        for a in QUA:
            s=list(a.attrib.values())
            L_QUA.append(str(s[0])+' '+str(s[1]))
        return L_QUA
    else:
        if 'QUANTITATIVE' not in keypoint:
            keypoint.append('QUANTITATIVE')
        return []
    
def INCOND1(INCOND,keypoint):
    L_INCOND=[]
    if INCOND:
        if 'INCOND' not in keypoint:
            keypoint.append('INCOND')

        for a in INCOND:
            s=list(a.attrib.values())
            L_INCOND.append(str(s[0])+' '+str(s[1])+" "+str(s[2]))
        return L_INCOND	
    else:
        if 'INCOND' not in keypoint:
            keypoint.append('INCOND')
        return [] 

def OUTCOND1(OUTCOND,keypoint):
    L_OUT=[]
    if OUTCOND:
        if 'OUTCOND' not in keypoint:
            keypoint.append('OUTCOND')
        for a in OUTCOND:
            s=list(a.attrib.values())
            L_OUT.append(str(s[0])+' '+str(s[1])+" "+str(s[2]))
        return L_OUT
    else:
        if 'OUTCOND' not in keypoint:
            keypoint.append('OUTCOND')
        return []

def ON1(ON,keypoint):
    L_ON=[]
    if ON:
        if 'ON' not in keypoint:
            keypoint.append('ON')
        key=['CODE']
        st=''
        for p in ON:
            k=list(p.attrib.keys())
            v=list(p.attrib.values())
            for q in key:
                if q in k:
                    st +=str(v[k.index(q)])+" "
                DOACTION=p.findall('DOACTION')
                if DOACTION:
                    for aB in DOACTION:
                        l=list(aB.attrib.values())
                        st +=" ".join(l)+" "
                DOCOND=p.findall('DOCOND')
                if DOCOND:
                    for aB in DOCOND:
                        l=list(aB.attrib.values())
                        st +=" ".join(l)+" "
                DOEMAIL=p.findall('DOMAIL')
                if DOEMAIL:
                    for aB in DOEMAIL:
                        l=list(aB.attrib.values())
                        st +=" ".join(l)+" "  
                DOSHOUT=p.findall('DOSHOUT')
                if DOSHOUT:
                    for aB in DOSHOUT:
                        l=list(aB.attrib.values())
                        st +=" ".join(l)+" "  
        L_ON.append(st)
        return L_ON           
    else:
        if 'ON' not in keypoint:
            keypoint.append('ON')
        return []
def SHOUT1(SHOUT,keypoint):
    L_SHOUT=[]
    if SHOUT:
        if 'NOTFY BEFOR(0)/AFTER' not in keypoint:
            keypoint.append('NOTFY BEFOR(0)/AFTER')
        for a in SHOUT:
            s=list(a.attrib.values())
            L_SHOUT.append(" ".join(s))
        return L_SHOUT
    else:
        if 'NOTFY BEFOR(0)/AFTER' not in keypoint:
            keypoint.append('NOTFY BEFOR(0)/AFTER')
        return []

def xml(file_name):   
    parsexml=et.parse(file_name+'.xml')
    root=parsexml.getroot()   
    keypoint=['PARENT_FOLDER','FOLDER_NAME','DATACENTER','FOLDER_ORDER_METHOD','RUN_AS','APPLICATION','SUB_APPLICATION','DESCRIPTION','SITE_STANDARD_NAME','TIMEFROM','TIMETO','CYCLIC','INTERVAL','MAXWAIT','MAXRERUN','CONFIRM']
    vaa=[[ 'PARENT_FOLDER','FOLDER_NAME','DATACENTER',   'FOLDER_ORDER_METHOD', 'RUN_AS', 'APPLICATION', 'SUB_APPLICATION','DESCRIPTION', 'SITE_STANDARD_NAME', 'TIMEFROM', 'TIMETO', 'CYCLIC', 'INTERVAL', 'MAXWAIT', 'MAXRERUN', 'VARIABLE', 'RBC', 'CONTROL', 'INCOND', 'OUTCOND', 'ON', 'NOTFY_BEFOR(0)/AFTER(1)']]   
    folder_details=vaa    
    for r in root:
        k=list(r.attrib.keys())
        v=list(r.attrib.values())
    
        va=[]
        for q in keypoint[:15]:
            if q in k:
                va.append(v[k.index(q)])
            else:
                if q=='FOLDER_ORDER_METHOD':
                    va.append('None Manuval Order')
                else:
                    va.append('')
        VARIABLE=r.findall('VARIABLE')
        va.append(VARIABLE1(VARIABLE,keypoint)) 
        
        RBC=r.findall('RULE_BASED_CALENDAR')
        va.append(RBC1(RBC,keypoint))
        
        CONTROL=r.findall('CONTROL')
        va.append(CONTROL1(CONTROL,keypoint))
        
        INCOND=r.findall('INCOND')
        va.append(INCOND1(INCOND,keypoint))
        
        OUTCOND=r.findall('OUTCOND')
        va.append(OUTCOND1(OUTCOND,keypoint))
        
        ON=r.findall('ON')
        va.append(ON1(ON,keypoint))
        
        SHOUT=r.findall('SHOUT')
        va.append(SHOUT1(SHOUT,keypoint))
        
        if va not in vaa:
            vaa.append(va)
    keypoint=['PARENT_FOLDER','JOBNAME','DATACENTER','FOLDER_ORDER_METHOD','RUN_AS','APPLICATION','SUB_APPLICATION','DESCRIPTION','SITE_STANDARD_NAME','TIMEFROM','TIMETO','CYCLIC','INTERVAL','MAXWAIT','MAXRERUN','CONFIRM']
    for r1 in root:
        subfolder=r1.findall("SUB_FOLDER")
        for r in subfolder:
            k=list(r.attrib.keys())
            v=list(r.attrib.values())
            va=[]
            for q in keypoint[:15]:
                if q in k:
                    va.append(v[k.index(q)])
                else:
                    if q=='FOLDER_ORDER_METHOD':
                        va.append('None Manuval Order')
                    else:
                        va.append('')
            vaa.append(va)
            VARIABLE=r.findall('VARIABLE')
            va.append(VARIABLE1(VARIABLE,keypoint)) 
        
            RBC=r.findall('RULE_BASED_CALENDAR')
            va.append(RBC1(RBC,keypoint))
        
            CONTROL=r.findall('CONTROL')
            va.append(CONTROL1(CONTROL,keypoint))
        
            INCOND=r.findall('INCOND')
            va.append(INCOND1(INCOND,keypoint))
        
            OUTCOND=r.findall('OUTCOND')
            va.append(OUTCOND1(OUTCOND,keypoint))
        
            ON=r.findall('ON')
            va.append(ON1(ON,keypoint))
        
            SHOUT=r.findall('SHOUT')
            va.append(SHOUT1(SHOUT,keypoint))
        
            if va not in vaa:
                vaa.append(va)
         
    for r1 in root:
        subfolder=r1.findall("SUB_FOLDER")
        for j in subfolder:
            subfolder=j.findall("SUB_FOLDER")
            for r in subfolder:
                k=list(r.attrib.keys())
                v=list(r.attrib.values())
                va=[]
                for q in keypoint[:15]:
                    if q in k:
                        va.append(v[k.index(q)])
                    else:
                        if q=='FOLDER_ORDER_METHOD':
                            va.append('None Manuval Order')
                        else:
                            va.append('')
                vaa.append(va)
                VARIABLE=r.findall('VARIABLE')
                va.append(VARIABLE1(VARIABLE,keypoint)) 
            
                RBC=r.findall('RULE_BASED_CALENDAR')
                va.append(RBC1(RBC,keypoint))
            
                CONTROL=r.findall('CONTROL')
                va.append(CONTROL1(CONTROL,keypoint))
            
                INCOND=r.findall('INCOND')
                va.append(INCOND1(INCOND,keypoint))
            
                OUTCOND=r.findall('OUTCOND')
                va.append(OUTCOND1(OUTCOND,keypoint))
            
                ON=r.findall('ON')
                va.append(ON1(ON,keypoint))
            
                SHOUT=r.findall('SHOUT')
                va.append(SHOUT1(SHOUT,keypoint))
            
                if va not in vaa:
                    vaa.append(va)
                
                
    
    keypoint1=['JOBNAME','PARENT_FOLDER','APPL_TYPE','DESCRIPTION','TASKTYPE','CMDLINE','NODEID','RUN_AS','APPLICATION','SUB_APPLICATION','PRIORITY','CONFIRM','RETRO','DAYS','TIMEFROM','TIMETO','CYCLIC','MAXRERUN','INTERVAL','MAXWAIT']
    vaa1=[['JOBNAME', 'PARENT_FOLDER', 'APPL_TYPE', 'DESCRIPTION', 'TASKTYPE', 'CMDLINE', 'NODEID', 'RUN_AS', 'APPLICATION', 'SUB_APPLICATION', 'PRIORITY', 'CONFIRM', 'RETRO', 'DAYS', 'TIMEFROM', 'TIMETO', 'CYCLIC', 'MAXRERUN', 'INTERVAL', 'MAXWAIT', 'VARIABLE', 'RBC', 'CONTROL', 'QUANTITATIVE', 'INCOND', 'OUTCOND', 'ON', 'NOTFY_BEFOR(0)/AFTER']]# data of the headers
    job_details=vaa1
    for r in root.findall("./SMART_FOLDER/SUB_FOLDER/"):
        JOB=r.findall('JOB')
        #finding the all attribuets of job
        for j in JOB:
            k=list(j.attrib.keys())       
            v=list(j.attrib.values())
            va=[]
            for q in keypoint1[:20]:  #30 is length of header as for the first
                if q in k:
                    va.append(v[k.index(q)])
                else:
                    va.append('')
            VARIABLE=j.findall('VARIABLE')
            va.append(VARIABLE1(VARIABLE,keypoint1)) 
                
            RBC=j.findall('RULE_BASED_CALENDARS')
            va.append(RBC1(RBC,keypoint1))
        
            CONTROL=j.findall('CONTROL')
            va.append(CONTROL1(CONTROL,keypoint1))
            
            QUA=j.findall('QUANTITATIVE')
            va.append(QUANTITATIVE1(QUA,keypoint1))
                
            INCOND=j.findall('INCOND')
            va.append(INCOND1(INCOND,keypoint1))
                
            OUTCOND=j.findall('OUTCOND')
            va.append(OUTCOND1(OUTCOND,keypoint1))
                
            ON=j.findall('ON')
            va.append(ON1(ON,keypoint1))
                
            SHOUT=j.findall('SHOUT')
            va.append(SHOUT1(SHOUT,keypoint1))
            if va not in vaa:
                vaa1.append(va)

    for r in root.findall("./SMART_FOLDER/"):
        JOB=r.findall('JOB')
        #finding the all attribuets of job
        for j in JOB:
            k=list(j.attrib.keys())       
            v=list(j.attrib.values())
            va=[]
            for q in keypoint1[:20]:  #30 is length of header as for the first
                if q in k:
                    va.append(v[k.index(q)])
                else:
                    va.append('')
            VARIABLE=j.findall('VARIABLE')
            va.append(VARIABLE1(VARIABLE,keypoint1)) 
                
            RBC=j.findall('RULE_BASED_CALENDARS')
            va.append(RBC1(RBC,keypoint1))
        
            CONTROL=j.findall('CONTROL')
            va.append(CONTROL1(CONTROL,keypoint1))
            
            QUA=j.findall('QUANTITATIVE')
            va.append(QUANTITATIVE1(QUA,keypoint1))
                
            INCOND=j.findall('INCOND')
            va.append(INCOND1(INCOND,keypoint1))
                
            OUTCOND=j.findall('OUTCOND')
            va.append(OUTCOND1(OUTCOND,keypoint1))
                
            ON=j.findall('ON')
            va.append(ON1(ON,keypoint1))
                
            SHOUT=j.findall('SHOUT')
            va.append(SHOUT1(SHOUT,keypoint1))
            if va not in vaa:
                vaa1.append(va)
    
    
    
    for r in root:
        JOB=r.findall('JOB')
        #finding the all attribuets of job
        for j in JOB:
            k=list(j.attrib.keys())       
            v=list(j.attrib.values())
        
            va=[]
            for q in keypoint1[:20]:  #30 is length of header as for the first
                if q in k:
                    va.append(v[k.index(q)])
                else:
                    va.append('')
          # variable keyword details appending in the data
            VARIABLE=j.findall('VARIABLE')
            va.append(VARIABLE1(VARIABLE,keypoint1)) 
                
            RBC=j.findall('RULE_BASED_CALENDARS')
            va.append(RBC1(RBC,keypoint1))
        
            CONTROL=r.findall('CONTROL')
            va.append(CONTROL1(CONTROL,keypoint1))
            
            QUA=j.findall('QUANTITATIVE')
            va.append(QUANTITATIVE1(QUA,keypoint1))
                
            INCOND=j.findall('INCOND')
            va.append(INCOND1(INCOND,keypoint1))
                
            OUTCOND=j.findall('OUTCOND')
            va.append(OUTCOND1(OUTCOND,keypoint1))
                
            ON=j.findall('ON')
            va.append(ON1(ON,keypoint1))
                
            SHOUT=j.findall('SHOUT')
            va.append(SHOUT1(SHOUT,keypoint1))
            if va not in vaa:
                vaa1.append(va)
    return folder_details,job_details

old="PRE"
new="POST"
output='CHANGES.xlsx'
def data(sheet,header,df1,df2):
    data=[list(df1.columns)]
    for c_index,col in enumerate(list(df1[header])):
        rows=df2[df2[header]==col].index.tolist() 
        if len(rows)>0:
            d1=list(df1.iloc[c_index])
            d2=list(df2.iloc[rows[0]])
            local=[]
            def new1(matches,addition,deletion,duplicates):
                if len(addition)>1 and len(deletion)>1:
                    new=matches+addition+deletion
                    if len(duplicates)>1:
                        new +=duplicates  
                    return new
                elif len(addition)>1 and len(deletion)<2:
                    new=matches+addition
                    if len(duplicates)>1:
                        new +=duplicates 
                    return new
                elif len(addition)<2 and len(deletion)>1:
                    new=matches+deletion
                    if len(duplicates)>1:
                        new +=duplicates
                    return new
                else:
                    if len(duplicates)>1:
                        return matches+duplicates 
                    else:
                        return matches
                
            for i,v in enumerate(d2):
                matches=[]
                addition=['addition']
                deletion=['deletion']
                duplicates=["duplicates"]
                if str(d1[i])==str(d2[i]) and type(v) != list:
                    local.append(d2[i])
                elif str(d1[i])!=str(d2[i]) and type(v) != list:
                    local.append("old: "+str(d1[i])+" \nnew: "+str(d2[i]))
                else:
                    if len(d2[i])==len(d1[i]):
                        d2_duplicate=[item for item, count in collections.Counter(d2[i]).items() if count > 1]
                        d1_duplicate=[item for item, count in collections.Counter(d1[i]).items() if count > 1]
                        dd=list(set(d1_duplicate+d2_duplicate))
                        if len(dd)>0:
                            duplicates.append(dd)
                        for da in v:
                            if da in d1[i]:
                                if da not in matches:
                                    matches.append(da)
                            else:
                                if da not in addition:
                                    addition.append(da)
                        for da in d1[i]:
                            if da in v:
                                if da not in matches:
                                    matches.append(da)
                            else:
                                if da not in deletion:
                                    deletion.append(da)    
                    elif len(d2[i])>len(d1[i]):
                        d2_duplicate=[item for item, count in collections.Counter(d2[i]).items() if count > 1]
                        d1_duplicate=[item for item, count in collections.Counter(d1[i]).items() if count > 1]
                        dd=list(set(d1_duplicate+d2_duplicate))
                        if len(dd)>0:
                            duplicates.append(dd)
                        for da in v:
                            if da in d1[i]:
                                if da not in matches:
                                    matches.append(da)
                            else:
                                if da not in addition:
                                    addition.append(da)    
                    else:
                        d2_duplicate=[item for item, count in collections.Counter(d2[i]).items() if count > 1]
                        d1_duplicate=[item for item, count in collections.Counter(d1[i]).items() if count > 1]
                        dd=list(set(d1_duplicate+d2_duplicate))
                        if len(dd)>0:
                            duplicates.append(dd)
                        for da in d1[i]:
                            if da in v:
                                if da not in matches:
                                    matches.append(da)
                            else:
                                if da not in deletion:
                                    deletion.append(da)
                    local.append(new1(matches,addition,deletion,duplicates))
                    

                                
            data.append(local)
    return data
data_old=xml(old)
folder,job=data_old[0],data_old[1]
df1_old = pd.DataFrame(columns=folder[0], data=folder[1:]).fillna("na_")
df2_old = pd.DataFrame(columns=job[0], data=job[1:]).fillna("na_")
print("old details found")
df2_old['PARENT_JOBNAME']=df2_old['PARENT_FOLDER']+df2_old['JOBNAME']

data_new=xml(new)
folder,job=data_new[0],data_new[1]
df1_new = pd.DataFrame(columns=folder[0], data=folder[1:])
df2_new = pd.DataFrame(columns=job[0], data=job[1:])
print('new data details found')
df2_new['PARENT_JOBNAME']=df2_new['PARENT_FOLDER']+df2_new['JOBNAME']


data_FOLDER=data('FOLDER','FOLDER_NAME',df1_old,df1_new)
new_folder = pd.DataFrame(columns=data_FOLDER[0], data=data_FOLDER[1:])


data_JOB=data('JOB','PARENT_JOBNAME',df2_old,df2_new) 
new_JOB = pd.DataFrame(columns=data_JOB[0], data=data_JOB[1:])

def deleted(df1,df2,column):
    deleted_folder_name=[]
    deleted_folder_data=[]
    for i in list(df1[column]):
        if i not in list(df2[column]):
            deleted_folder_name.append(i)
            deleted_folder_data.append(list(df1.iloc[df1[df1[column]==i].index.tolist()[0]]))
    return deleted_folder_name,deleted_folder_data
deleted_folder=deleted(df1_old,df1_new,'FOLDER_NAME')
deleted_JOB=deleted(df2_old,df2_new,'PARENT_JOBNAME')
cel_deleted_job=[]
for i in deleted_JOB[1]:
    new_JOB.loc[len(new_JOB)] = i
    c=list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')+['AA','AB']    
    for j in c:
        cel_deleted_job.append(j+str(len(new_JOB)+1))
cel_deleted_Folder=[]
for i in deleted_folder[1]:
    new_folder.loc[len(new_folder)] = i
    c=list('ABCDEFGHIJKLMNOPQRSTUV')    
    for j in c:
        cel_deleted_Folder.append(j+str(len(new_folder)+1))
        
def addition(df1,df2,column):
    addition_folder_name=[]
    addition_folder_data=[]
    for i in list(df2[column]):
        if i not in list(df1[column]):
            addition_folder_name.append(i)
            addition_folder_data.append(list(df2.iloc[df2[df2[column]==i].index.tolist()[0]]))
    return addition_folder_name,addition_folder_data

addition_folder=addition(df1_old,df1_new,'FOLDER_NAME')
addition_JOB=addition(df2_old,df2_new,'PARENT_JOBNAME')

cel_addition_job=[]
for i in addition_JOB[1]:
    new_JOB.loc[len(new_JOB)] = i
    c=list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')+['AA','AB']    
    for j in c:
        cel_addition_job.append(j+str(len(new_JOB)+1))
cel_addition_Folder=[]
for i in addition_folder[1]:
    new_folder.loc[len(new_folder)] = i
    c=list('ABCDEFGHIJKLMNOPQRSTUV')    
    for j in c:
        cel_addition_Folder.append(j+str(len(new_folder)+1))
new_folder['RBC']=["\n".join(i) if len(i)!=0 else "" for i in new_folder['RBC']]
new_folder['VARIABLE']=["\n".join(i) if len(i)!=0 else "" for i in new_folder['VARIABLE']]
new_folder['INCOND']=["\n".join(i) for i in new_folder['INCOND']]
new_folder['OUTCOND']=["\n".join(i) for i in new_folder['OUTCOND']]
new_folder['ON']=["\n".join(i) for i in new_folder['ON']]
new_folder['NOTFY_BEFOR(0)/AFTER(1)']=["\n".join(i) for i in new_folder['NOTFY_BEFOR(0)/AFTER(1)']]
new_JOB['INCOND']=["\n".join(i) for i in new_JOB['INCOND']]
new_JOB['VARIABLE']=["\n".join(i) if len(i)!=0 else "" for i in new_JOB['VARIABLE']]
new_JOB['OUTCOND']=["\n".join(i) for i in new_JOB['OUTCOND']]
new_JOB['ON']=["\n".join(i) for i in new_JOB['ON']]
new_JOB['NOTFY_BEFOR(0)/AFTER']=["\n".join(i) for i in new_JOB['NOTFY_BEFOR(0)/AFTER']]

with pd.ExcelWriter(output) as writer:
    new_folder.iloc[:,:-1].to_excel(writer, sheet_name='FOLDER',index=False)
    new_JOB.iloc[:,:-1].to_excel(writer, sheet_name='JOB',index=False)

    

import openpyxl
wb = openpyxl.load_workbook(output)
def fill_color(output,sheet,cel_deleted,cel_addition):
    dataframe=pd.read_excel(output, sheet_name=sheet)
    ws = wb[sheet]
    fill_cell6 = PatternFill(patternType='solid', fgColor='99CC32')
    for k in cel_addition:
        ws[k].fill = fill_cell6
        
    fill_cell5 = PatternFill(patternType='solid', fgColor='EE5C42')
    for j in cel_deleted:
        ws[j].fill = fill_cell5
    
    fill_cell4 = PatternFill(patternType='solid', fgColor='FFEC8B')
    fill_cell5 = PatternFill(patternType='solid', fgColor='FFD39B')
    for i,value in enumerate(list(dataframe.columns)):
        for j,v in enumerate(list(dataframe[value])):
            if 'old:' in str(v) or 'addition' in str(v) or 'deletion' in str(v) or 'duplicates' in str(v):
                if i>25:
                    c='A'+chr(i-26+65)+str(j+2)
                    ws[c].fill = fill_cell4
                else:
                    c=chr(i+65)+str(j+2)
                    ws['A'+str(j+2)].fill = fill_cell5
                    ws[c].fill = fill_cell4
fill_color(output,'JOB',cel_deleted_job,cel_addition_job)
fill_color(output,'FOLDER',cel_deleted_Folder,cel_addition_Folder)

wb.save(output)
print('succuss')